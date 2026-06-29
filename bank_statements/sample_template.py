"""Downloadable Excel sample template for the Analee bank-statement upload.

Returns an in-memory ``.xlsx`` (BytesIO) with:
  * a **data sheet** — the exact header row the reader expects + realistic,
    valid sample rows; and
  * an **Instructions sheet** — accepted formats and what is **NOT** accepted.

Built with openpyxl (already a dependency; Analee has no xlsxwriter).
``BANK_STATEMENT_COLUMNS`` mirrors ``BankStatementExcelReader.required_columns``
and is locked to it by the guard tests in
``tests/test_bank_statement_template.py``.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

# Mirrors bank_statements/excel_reader.py: self.required_columns
BANK_STATEMENT_COLUMNS = ['Date', 'Description', 'Amount']

# Amount is a single signed number: positive = money IN, negative = money OUT.
_SAMPLE_ROWS = [
    ['2026-01-05', 'Salary deposit', 25000],
    ['2026-01-08', 'Office rent', -7500],
    ['2026-01-15', 'Client payment received', 12000],
    ['2026-01-20', 'Bank charges', -150.50],
]

_INSTRUCTIONS = [
    'Analee — Bank Statement upload template',
    '',
    'HOW TO USE: enter your transactions on the "Bank Statement" sheet under the '
    'headers provided, then upload the file (.xlsx or .csv).',
    '',
    'COLUMNS (case-insensitive; column order does not matter; extra columns are ignored):',
    '  • Date — the transaction date. Any common format works (2026-01-31 or 31/01/2026).',
    '  • Description — what the transaction was (up to 200 characters).',
    '  • Amount — a single signed number: POSITIVE = money IN, NEGATIVE = money OUT. '
    'Currency symbols and thousands commas are removed automatically (e.g. R1,200.00).',
    '',
    'OPTIONAL: a "Category" column may be included; it is tolerated if present.',
    '',
    'WHAT IS NOT ACCEPTED:',
    '  • A future date (a transaction dated after today).',
    '  • A blank Description, or a file with no data rows.',
    '  • A statement spanning more than one year (366 days) between its earliest and latest date.',
]


def build_bank_statement_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bank Statement'

    ws.append(BANK_STATEMENT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in _SAMPLE_ROWS:
        ws.append(row)
    for col in ('A', 'B', 'C'):
        ws.column_dimensions[col].width = 26

    notes = wb.create_sheet('Instructions')
    notes.column_dimensions['A'].width = 110
    for line in _INSTRUCTIONS:
        notes.append([line])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
