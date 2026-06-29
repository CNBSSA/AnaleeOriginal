"""Trial balance export — BooksXperts-compatible Excel."""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from models import Account, CompanySettings, Transaction, db
from reports.trial_balance_service import (
    BOOKSXPERTS_TB_COLUMNS,
    TrialBalanceRow,
    build_booksxperts_trial_balance_xlsx,
    export_filename,
    load_trial_balance,
)


def _settings(user_id: int, fy_end: int = 2) -> CompanySettings:
    settings = CompanySettings(
        user_id=user_id,
        company_name='ACME Pty Ltd',
        financial_year_end=fy_end,
    )
    db.session.add(settings)
    db.session.commit()
    return settings


def test_build_xlsx_uses_booksxperts_columns():
    rows = (
        TrialBalanceRow('ca.810.001', 'Bank Cheque Account 1', Decimal('50000')),
        TrialBalanceRow('i.100.000', 'Sales', Decimal('-50000')),
    )
    data = build_booksxperts_trial_balance_xlsx(
        rows,
        company_name='Test Co',
        period_end=datetime(2026, 2, 28),
    )
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    assert list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] == BOOKSXPERTS_TB_COLUMNS
    data_rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
    assert data_rows[0] == ('ca.810.001', 'Bank Cheque Account 1', 50000.0)
    assert data_rows[1] == ('i.100.000', 'Sales', -50000.0)
    footer_rows = [row[0] for row in ws.iter_rows(min_row=5, values_only=True) if row[0]]
    footer_text = ' '.join(str(cell) for cell in footer_rows)
    assert 'BooksXperts' in footer_text
    assert 'Accountants' in footer_text


def test_export_filename_includes_period_end():
    end = datetime(2026, 2, 28)
    assert export_filename(end) == 'analee-trial-balance-2026-02-28.xlsx'


def test_load_trial_balance_builds_signed_amounts(app, sample_user):
    with app.app_context():
        _settings(sample_user)
        bank = Account(
            link='ca.810.001',
            name='Bank Cheque Account 1',
            category='Assets',
            sub_category='Current Asset',
            user_id=sample_user,
        )
        sales = Account(
            link='i.100.000',
            name='Sales',
            category='Income',
            sub_category='Income',
            user_id=sample_user,
        )
        db.session.add_all([bank, sales])
        db.session.flush()
        db.session.add_all([
            Transaction(
                date=datetime(2026, 4, 15),
                description='Receipt',
                amount=100.0,
                user_id=sample_user,
                account_id=bank.id,
            ),
            Transaction(
                date=datetime(2026, 4, 16),
                description='Sale',
                amount=-100.0,
                user_id=sample_user,
                account_id=sales.id,
            ),
        ])
        db.session.commit()

        ctx = load_trial_balance(sample_user)
        assert len(ctx.rows) == 2
        by_link = {r.link: r.amount for r in ctx.rows}
        assert by_link['ca.810.001'] == Decimal('100.00')
        assert by_link['i.100.000'] == Decimal('-100.00')
        assert ctx.total_debits == ctx.total_credits == Decimal('100.00')


def test_load_trial_balance_skips_zero_balances(app, sample_user):
    with app.app_context():
        _settings(sample_user)
        empty = Account(
            link='e.460.000',
            name='Salaries',
            category='Expenses',
            sub_category='Expense',
            user_id=sample_user,
        )
        db.session.add(empty)
        db.session.commit()

        ctx = load_trial_balance(sample_user)
        assert ctx.rows == ()
