"""Guard tests: the downloadable bank-statement template must stay in sync with
the real reader (bank_statements/excel_reader.py). If the reader's required
columns change, these tests fail until the template is updated — so a user can
never download a template that then fails to import.
"""
import os
from datetime import datetime

from openpyxl import load_workbook

from bank_statements.sample_template import (
    BANK_STATEMENT_COLUMNS,
    build_bank_statement_template,
)

_READER_SRC = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), 'bank_statements', 'excel_reader.py')


def _rows():
    wb = load_workbook(build_bank_statement_template(), read_only=True, data_only=True)
    ws = wb['Bank Statement']
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def test_headers_match_reader():
    rows = _rows()
    assert [str(c).strip() for c in rows[0]] == BANK_STATEMENT_COLUMNS


def test_sample_rows_are_valid_per_reader_rules():
    rows = _rows()
    assert len(rows) > 1, 'template should ship with sample rows'
    date_i = BANK_STATEMENT_COLUMNS.index('Date')
    desc_i = BANK_STATEMENT_COLUMNS.index('Description')
    amount_i = BANK_STATEMENT_COLUMNS.index('Amount')

    dates = []
    for row in rows[1:]:
        assert isinstance(row[amount_i], (int, float)), f'amount must be numeric: {row}'
        desc = str(row[desc_i]).strip()
        assert 0 < len(desc) <= 200, f'description must be 1..200 chars: {desc!r}'
        parsed = datetime.strptime(str(row[date_i]), '%Y-%m-%d')
        assert parsed <= datetime.now(), f'sample date must not be in the future: {row[date_i]}'
        dates.append(parsed)

    span_days = (max(dates) - min(dates)).days
    assert span_days <= 366, f'sample statement must span <= 366 days, got {span_days}'


def test_columns_locked_to_reader_source():
    with open(_READER_SRC, encoding='utf-8') as fh:
        src = fh.read()
    for col in BANK_STATEMENT_COLUMNS:
        assert f"'{col}'" in src, f'reader no longer requires column "{col}"'
